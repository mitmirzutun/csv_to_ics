import csv_ical
import abc
import typing
import os
import datetime
import csv
convert = csv_ical.Convert()


class Event:
    def __init__(self, event_id: str, title: str, summary: str, description: str,
                 location: str, start: datetime.datetime,
                 stop: datetime.datetime,
                 event_class: typing.Literal["PUBLIC"]):
        self.__id = event_id
        self.__location = location
        self.__summary = summary
        self.__description = description
        self.__title = title
        self.__start = start.astimezone(datetime.timezone.utc)
        self.__stop = stop.astimezone(datetime.timezone.utc)
        self.__event_class = event_class
        self.__created = datetime.datetime.now()

    def to_list(self) -> typing.List[typing.Any]:
        return [self.__id, self.__location, self.__summary,
                self.__description, self.__title, self.__start, self.__stop,
                self.__created, self.__event_class]

    def to_ics_event(self) -> str:
        return ("BEGIN:VEVENT\n" +
                f"UID:{self.__id}\n" +
                f"LOCATION:{self.__location}\n" +
                f"SUMMARY:{self.__summary}\n" +
                f"DESCRIPTION:{self.__description}\n" +
                "DTSTART:{}\n".format(self.__start
                                      .strftime("%Y%m%dT%H%M%SZ")) +
                "DTEND:{}\n".format(self.__stop.strftime("%Y%m%dT%H%M%SZ")) +
                f"DTSTAMP:{self.__created}\n" +
                f"STATUS:{self.__event_class}\n" +
                "END:VEVENT\n"
                )


class Calender:
    def __init__(self, events: list[Event] | Event):
        self.__events = events if isinstance(events, list) else [events]

    def to_ical(self, filename: str):
        fd = open(filename, "w")
        fd.write("BEGIN:VCALENDAR\n")
        fd.write("VERSION:2.0\n")
        fd.write("PRODID:NULL\n")
        for event in self.__events:
            fd.write(event.to_ics_event())
        fd.write("END:VCALENDAR")

    def to_csv(self, filename: str):
        import csv
        writer = csv.writer(open(filename, "w"),)
        writer.writerow(["ID", "Location", "Summary", "Description", "Title",
                         "Start", "Stop", "Created", "EventClass"])
        for event in self.__events:
            writer.writerow(event.to_list())


class CalenderBuilder(abc.ABC):
    @abc.abstractmethod
    def from_path(self, path: str) -> typing.Optional[Calender
                                                      | list[Calender]]:
        return NotImplemented


class OVBuilder(CalenderBuilder):
    def __init__(self):
        pass
    
    def from_path(self, path: str) -> typing.Optional[Calender]:
        with open(path, "r", newline="") as fd:
            reader = csv.reader(fd, delimiter=",", lineterminator="\n")
            events = []
            next(reader)
            for row in reader:
                location = row[5]
                summary = row[1]
                print(row[0]+" "+row[2])
                try:
                    start = datetime.datetime.strptime(row[0]+" "+row[2],
                                                       "%d.%m.%Y %H:%M")
                    stop = datetime.datetime.strptime(row[0]+" "+row[3],
                                                      "%d.%m.%Y %H:%M")
                except ValueError:
                    print("Date, start time or end time missing: {}".format(row
                                                                            ))
                    continue
                description = row[10]
                event_id = summary
                title = summary
                tmp = Event(event_id, title, summary, description, location, start,
                            stop, "PUBLIC")
                events.append(tmp)
        return Calender(events)


class CSVBuilder(CalenderBuilder):
    def __init__(self) -> None:
        self.__delimiter: str = ","
        self.__location: typing.Optional[int] = None
        self.__summary: typing.Optional[int] = None
        self.__description: typing.Optional[int] = None
        self.__title: typing.Optional[int] = None
        self.__start_date: typing.Optional[int] = None
        self.__stop_date: typing.Optional[int] = None
        self.__start_time: typing.Optional[int] = None
        self.__stop_time: typing.Optional[int] = None
        self.__num_headers: typing.Optional[int] = None

    def location_col(self, col: int) -> 'CSVBuilder':
        self.__location = col
        return self

    def title_col(self, col: int) -> 'CSVBuilder':
        self.__title = col
        return self

    def summary_col(self, col: int) -> 'CSVBuilder':
        self.__summary = col
        if self.__title is None:
            self.__title = col
        return self

    def description_col(self, col: int) -> 'CSVBuilder':
        self.__description = col
        return self

    def start_date_col(self, col: int) -> 'CSVBuilder':
        self.__start_date = col
        return self

    def end_date_col(self, col: int) -> 'CSVBuilder':
        self.stop_date = col
        return self

    def start_time_col(self, col: int) -> 'CSVBuilder':
        self.__start_time = col
        return self

    def end_time_col(self, col: int) -> 'CSVBuilder':
        self.__stop_time = col
        return self

    def number_of_headers(self, number: int) -> 'CSVBuilder':
        self.__num_headers = number
        return self

    def from_path(self, path: str) -> typing.Optional[Calender]:
        if self.__start_time is None:
            return None
        if self.__stop_time is None:
            return None
        location_row = self.__location
        num_headers: int
        if self.__num_headers is None:
            num_headers = 0
        else:
            num_headers = self.__num_headers
        import csv
        file_descriptor = open(path, "r")
        reader = csv.reader(file_descriptor, delimiter=self.__delimiter)
        for _ in range(num_headers):
            next(reader)
        events = []
        for row in reader:
            location: str
            if self.__location is None:
                location = ""
            else:
                location = row[location_row]
            date: str
            if self.__start_time is None:
                date = row[self.__start_date]
            else:
                date = row[self.__start_date]+" "+row[self.__start_time]
            start = datetime.datetime.strptime(date,
                                               "%Y-%m-%d %H:%M:%S")
            if self.__stop_date is None:
                date = row[self.__stop_date]
            else:
                date = row[self.__stop_date]+" "+row[self.__stop_time]
            stop = datetime.datetime.strptime(date,
                                              "%Y-%m-%d %H:%M:%S")
            title: str
            if self.__title is None:
                title = ""
            else:
                title = row[self.__title]
            summary: str
            if self.__summary is None:
                summary = ""
            else:
                summary = row[self.__summary]
            description: str
            if self.__description is None:
                description = ""
            else:
                description = row[self.__description]
            tmp = Event(title, title, summary, description, location, start,
                        stop, "PUBLIC")
            events.append(tmp)
        file_descriptor.close()
        return Calender(events)


def ics_to_csv(ics_file_name, csv_file_name):
    convert.read_ical(ics_file_name)
    convert.make_csv()
    convert.save_csv(csv_file_name)


def xls_to_ics(xlsx_file_name, ics_file_name):
    import openpyxl
    import csv
    import datetime
    xlsx_file = openpyxl.load_workbook(xlsx_file_name)
    worksheet = xlsx_file.active
    with open(ics_file_name + ".tmp.csv", "w") as csv_file:
        writer = csv.writer(csv_file)
        title_row = next(worksheet.rows)
        titles = list(val.value for val in title_row)
        print(title_row, titles)
        XLSX_BOOKING_ID: int = titles.index("bookingCode") + 1
        XLSX_INTERNAL: int = titles.index("price.groupInternal") + 1
        XLSX_STATUS: int = titles.index("status") + 1
        XLSX_PERSONS: int = titles.index("persons") + 1
        XLSX_RESOURCE: int = titles.index("resourceTitle") + 1
        XLSX_START_DATE: int = titles.index("bookingStartAtDate") + 1
        XLSX_START_TIME: int = titles.index("bookingStartAtTime") + 1
        XLSX_END_DATE: int = titles.index("bookingEndAtDate") + 1
        XLSX_END_TIME: int = titles.index("bookingEndAtTime") + 1
        XLSX_DESCRIPTION: int = titles.index("customerNote") + 1
        print(worksheet.cell(column=XLSX_BOOKING_ID, row=1).value)
        for row in range(2, worksheet.max_row + 1):
            print(worksheet.cell(column=XLSX_BOOKING_ID, row=row).value)
            if (worksheet.cell(column=XLSX_STATUS, row=row).value !=
                    "completed"):
                continue
            if (worksheet.cell(column=XLSX_RESOURCE, row=row).value !=
                    "Pavillon Gemeinschaftsraum - Kooperative Großstadt "):
                continue
            booking_id = worksheet.cell(column=XLSX_BOOKING_ID, row=row).value
            if worksheet.cell(column=XLSX_INTERNAL, row=row).value:
                booking_id += " (INTERNAL)"
            else:
                booking_id += " (EXTERNAL)"
            persons = worksheet.cell(column=XLSX_PERSONS, row=row).value
            start_date = worksheet.cell(column=XLSX_START_DATE, row=row
                                        ).value + " " +\
                worksheet.cell(column=XLSX_START_TIME, row=row).value
            end_date = worksheet.cell(column=XLSX_END_DATE, row=row).value +\
                " " +\
                worksheet.cell(column=XLSX_END_TIME, row=row).value
            desc = worksheet.cell(column=XLSX_DESCRIPTION, row=row).value
            writer.writerow([booking_id, start_date, end_date, desc,
                             "Freihampton", persons])
    config = {
        'CSV_NAME': 0,
        'CSV_START_DATE': 1,
        'CSV_END_DATE': 2,
        'CSV_DESCRIPTION': 3,
        'CSV_LOCATION': 4,
        'CSV_PERSONS': 5,
        'CSV_DELIMITER': ","
    }
    convert.read_csv(ics_file_name + ".tmp.csv", config)
    for i in range(len(convert.csv_data)):
        tmp = (datetime.datetime
               .strptime(convert.csv_data[i][config["CSV_START_DATE"]],
                         "%d.%m.%Y %H:%M"))
        convert.csv_data[i][config["CSV_START_DATE"]] = tmp
        tmp = datetime.datetime.strptime(convert
                                         .csv_data[i][config["CSV_END_DATE"]],
                                         "%d.%m.%Y %H:%M")
        convert.csv_data[i][config["CSV_END_DATE"]] = tmp
    convert.make_ical(config)
    convert.save_ical(ics_file_name)
    os.remove(ics_file_name + ".tmp.csv")
