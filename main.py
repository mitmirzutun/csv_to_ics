import csv_ical
convert=csv_ical.Convert()
def ics_to_csv(ics_file_name,csv_file_name):
    convert.read_ical(ics_file_name)
    convert.make_csv()
    convert.save_csv(csv_file_name)
def xls_to_ics(xlsx_file_name,ics_file_name):
    import openpyxl,csv,datetime
    xlsx_file=openpyxl.load_workbook(xlsx_file_name)
    worksheet=xlsx_file.active
    with open(ics_file_name+".tmp.csv","w") as csv_file:
        writer=csv.writer(csv_file)
        for row in range(2,worksheet.max_row+1):
            print(worksheet.cell(column=1,row=row).value)
            if worksheet.cell(column=17,row=row).value!="completed":
                continue
            booking_id=worksheet.cell(column=1,row=row).value
            start_date=worksheet.cell(column=21,row=row).value+" "+worksheet.cell(column=22,row=row).value
            end_date=worksheet.cell(column=23,row=row).value+" "+worksheet.cell(column=24,row=row).value
            desc=worksheet.cell(column=31,row=row).value
            writer.writerow([booking_id,start_date,end_date,desc,"Freihampton"])
    config={
        'CSV_NAME':0,
        'CSV_START_DATE':1,
        'CSV_END_DATE':2,
        'CSV_DESCRIPTION':3,
        'CSV_LOCATION':4,
        'CSV_DELIMITER':","
    }
    convert.read_csv(ics_file_name+".tmp.csv",config)
    for i in range(len(convert.csv_data)):
        convert.csv_data[i][config["CSV_START_DATE"]]=datetime.datetime.strptime(convert.csv_data[i][config["CSV_START_DATE"]], "%Y-%m-%d %H:%M")
        convert.csv_data[i][config["CSV_END_DATE"]]=datetime.datetime.strptime(convert.csv_data[i][config["CSV_END_DATE"]], "%Y-%m-%d %H:%M")
    convert.make_ical(config)
    convert.save_ical(ics_file_name)
def main():
    file_name=input("File name: ")
    if file_name.endswith(".ics"):
        ics_to_csv(file_name,input("CSV file name: "))
    if file_name.endswith(".xlsx"):
        xls_to_ics(file_name,input("ICS file name: "))
if __name__=="__main__":
    main()