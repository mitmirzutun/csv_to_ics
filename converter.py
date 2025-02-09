import csv_ical
import os
convert = csv_ical.Convert()


def ics_to_csv(ics_file_name, csv_file_name):
    convert.read_ical(ics_file_name)
    convert.make_csv()
    convert.save_csv(csv_file_name)


def xls_to_ics(xlsx_file_name, ics_file_name):
    import openpyxl
    import csv
    import datetime
    xlsx_file = openpyxl.load_workbook(xlsx_file_name)
    worksheet = xlsx_file.active
    with open(ics_file_name + ".tmp.csv", "w") as csv_file:
        writer = csv.writer(csv_file)
        title_row = next(worksheet.rows)
        titles = list(val.value for val in title_row)
        print(title_row, titles)
        XLSX_BOOKING_ID: int = titles.index("bookingCode") + 1
        XLSX_INTERNAL: int = titles.index("price.groupInternal") + 1
        XLSX_STATUS: int = titles.index("status") + 1
        XLSX_PERSONS: int = titles.index("persons") + 1
        XLSX_RESOURCE: int = titles.index("resourceTitle") + 1
        XLSX_START_DATE: int = titles.index("bookingStartAtDate") + 1
        XLSX_START_TIME: int = titles.index("bookingStartAtTime") + 1
        XLSX_END_DATE: int = titles.index("bookingEndAtDate") + 1
        XLSX_END_TIME: int = titles.index("bookingEndAtTime") + 1
        XLSX_DESCRIPTION: int = titles.index("customerNote") + 1
        print(worksheet.cell(column=XLSX_BOOKING_ID, row=1).value)
        for row in range(2, worksheet.max_row + 1):
            print(worksheet.cell(column=XLSX_BOOKING_ID, row=row).value)
            if worksheet.cell(column=XLSX_STATUS, row=row).value != "completed":
                continue
            if worksheet.cell(column=XLSX_RESOURCE, row=row).value != "Pavillon Gemeinschaftsraum - Kooperative Großstadt ":
                continue
            booking_id = worksheet.cell(column=XLSX_BOOKING_ID, row=row).value
            if worksheet.cell(column=XLSX_INTERNAL, row=row).value:
                booking_id += " (INTERNAL)"
            else:
                booking_id += " (EXTERNAL)"
            persons = worksheet.cell(column=XLSX_PERSONS, row=row).value
            start_date = worksheet.cell(column=XLSX_START_DATE, row=row).value + " " + worksheet.cell(column=XLSX_START_TIME, row=row).value
            end_date = worksheet.cell(column=XLSX_END_DATE, row=row).value + " " + worksheet.cell(column=XLSX_END_TIME, row=row).value
            desc = worksheet.cell(column=XLSX_DESCRIPTION, row=row).value
            writer.writerow([booking_id, start_date, end_date, desc, "Freihampton", persons])
    config = {
        'CSV_NAME': 0,
        'CSV_START_DATE': 1,
        'CSV_END_DATE': 2,
        'CSV_DESCRIPTION': 3,
        'CSV_LOCATION': 4,
        'CSV_PERSONS': 5,
        'CSV_DELIMITER': ","
    }
    convert.read_csv(ics_file_name + ".tmp.csv", config)
    for i in range(len(convert.csv_data)):
        convert.csv_data[i][config["CSV_START_DATE"]] = datetime.datetime.strptime(convert.csv_data[i][config["CSV_START_DATE"]], "%d.%m.%Y %H:%M")
        convert.csv_data[i][config["CSV_END_DATE"]] = datetime.datetime.strptime(convert.csv_data[i][config["CSV_END_DATE"]], "%d.%m.%Y %H:%M")
    convert.make_ical(config)
    convert.save_ical(ics_file_name)
    os.remove(ics_file_name + ".tmp.csv")
